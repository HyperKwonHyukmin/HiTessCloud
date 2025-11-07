import socket
import threading
import os
import subprocess
import time
import math
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Font
import sys


original_file_path = r"C:\Users\HHI\KHM\HiTessCloud_Flask\main\PythonModule\Reference\Report_Ladder.xlsx"
wb = load_workbook(filename=original_file_path)
sheet = wb["구조 강도 평가"]
print(sheet["b12"].value)
