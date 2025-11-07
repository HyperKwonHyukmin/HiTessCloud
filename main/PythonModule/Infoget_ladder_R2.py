import socket
import threading
import os
import subprocess
import time
import math
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

if len(sys.argv) < 2:
    print("Usage: infoget_ladder.py <파일경로>")
    sys.exit(1)

file_path = sys.argv[1]  # 첫 번째 인자: 파일 경로
file_name = os.path.basename(file_path)
dir_path = os.path.dirname(file_path)

# 2) 해석 폴더 생성 ──────────────────────────────────────────────────────────────────────────────

Analysis_bdf = file_path
Analysis_f06 = file_path.replace(".bdf",".f06")
Analysis_op2 = file_path.replace(".bdf",".op2")

# 3) 해석 파일 실행 ──────────────────────────────────────────────────────────────────────────────

os.chdir(dir_path)
subprocess.Popen('nastran ' + Analysis_bdf).wait()


# 4) 해석 파일 분석 ──────────────────────────────────────────────────────────────────────────────
with open(Analysis_bdf, 'r', encoding='utf8') as f:
    bdf = f.readlines()
with open(Analysis_f06, 'r', encoding='utf8') as f:
    f06 = f.readlines()

# 4-1) bdf 분석
'''와이어 안전 하중 List 생성'''
list_BL = [[8.0, 3.5],
           [10.0, 5.6],
           [12.0, 8.4],
           [14.0, 11.2],
           [16.0, 14.7],
           [18.0, 18.9],
           [20.0, 23.1],
           [22.0, 28.0],
           [24.0, 33.6]]
'''내용 List 생성'''
list_Info=[]
list_Set = []
list_Prop = []
list_Rod = []
collecting = False
temp_set = []
## BDF 파일에서 내용 추출 (SET, Prop, Rod)
for i, s in enumerate(bdf):
    # 5.1 BDF 파일에서 내용 추출 (SET, MAT, PROP)
    if s.find('$-') == 0:  # '$-'가 문자열의 시작
        name, value = s.replace("$-", "").split(",")
        name_parts = name.rsplit("_", 1)[0]
        parts = name_parts.split("_")
        comp_name = "_".join(parts[:-1])
        length = parts[-1]
        if "p" in length:
            length = length.replace("p", ".")
        name_clean = f"{comp_name}_{length}"
        list_Info.append([name_clean, int(value)])  # value를 숫자로 변환
    if s.strip().startswith("SET"):
        collecting = True
        # "SET 1 =" 잘라내고 오른쪽 값만
        parts = s.split("=", 1)[1]
        nums = parts.replace(",", " ").split()
        temp_set.extend(int(x) for x in nums)
        if not s.strip().endswith(","):  # ,로 안 끝나면 SET 끝
            list_Set.append(temp_set)
            temp_set = []
            collecting = False
    # SET 이어지는 라인
    elif collecting:
        nums = s.replace(",", " ").split()
        temp_set.extend(int(x) for x in nums)
        if not s.strip().endswith(","):
            list_Set.append(temp_set)
            temp_set = []
            collecting = False
    if s.find('PROD    ') != -1:
        Prop_No = s[8:16].strip()
        Mat_No = s[16:24].strip()
        Area = float(s[24:32].strip())
        Radius=math.sqrt(Area/math.pi)
        diameter = round(Radius, 1)*2
        prop_data = [int(Prop_No), int(Mat_No), float(Area),float(diameter)]
        list_Prop.append(prop_data)
    if s.find('CROD    ') != -1:
        Elem_No = s[8:16].strip()
        Prop_No = s[16:24].strip()
        Node1 = s[24:32].strip()
        Node2 = s[32:40].strip()
        rod_data = [int(Elem_No), int(Prop_No), int(Node1), int(Node2)]
        list_Rod.append(rod_data)

'''Dataframe 생성'''
df_Set = pd.DataFrame(list_Set[0], columns=["Elem_No"])
df_Prop = pd.DataFrame(list_Prop, columns=['Prop_No', 'Mat_No','Area','Dia'])
df_Rod = pd.DataFrame(list_Rod, columns=['Elem_No', 'Prop_No', 'Node1', 'Node2'])
df_Info = pd.DataFrame(list_Info, columns=["Comp", "Load"])
df_BL = pd.DataFrame(list_BL, columns=["Dia", "BL[ton]"])
df_Set = df_Set.merge(df_Rod[['Elem_No', 'Prop_No']], on='Elem_No', how='left')
df_Set = df_Set.merge(df_Prop[['Prop_No','Dia']], on='Prop_No', how='left')
df_Set = df_Set.merge(df_BL[['Dia', 'BL[ton]']], on='Dia', how='left')

# 4-2) f06 분석
# ① 변형에 해당 하는 Index 추출
dummy_disp_index, dummy_disp_index2 = [], []
for i, s in enumerate(f06):
    if s.find('D I S P L A C E M E N T   V E C T O R') != -1:
        dummy_disp_index.append(i)
        start_disp_index = min(dummy_disp_index) - 2
for i, s in enumerate(f06):
    if s.find('F O R C E S ') != -1:
        dummy_disp_index2.append(i)
        end_disp_index = min(dummy_disp_index2) - 4
displacement = f06[start_disp_index:end_disp_index]
# ② 변형 데이터 정리
del_disp_list = []
for i, s in enumerate(displacement):
    if s.find('D I S P L A C E M E N T   V E C T') != -1:
        del_disp_list.append(i)
del_disp_list.reverse()
for j in range(0, len(del_disp_list)):
    if j < len(del_disp_list) - 1:
        del displacement[del_disp_list[j] - 4:del_disp_list[j] + 3]
    if j == len(del_disp_list) - 1:
        del displacement[0:del_disp_list[j] + 3]
for j in range(0, len(displacement)):
    displacement[j] = displacement[j].split()
disp_final_list = []
for k in range(0, len(displacement) - 1):
    disp_final_list.append([int(displacement[k][0]), float(displacement[k][2]), float(displacement[k][3]), float(displacement[k][4]), (float(displacement[k][2]) ** 2 + float(displacement[k][3]) ** 2 + float(displacement[k][4]) ** 2) ** (1 / 2)])
df_disp = pd.DataFrame(disp_final_list, columns=['NODE', 'DISP_X', 'DISP_Y', 'DISP_Z', 'DISP'])
max_disp = df_disp['DISP'].max()
max_disp_element_no = df_disp.loc[df_disp['DISP'].idxmax(), 'NODE']

# ③ 응력에 해당 하는  Index 추출
dummy_stress_index, dummy_stress_index2 = [], []
for i, s in enumerate(f06):
    if s.find('S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M ') != -1:
        dummy_stress_index.append(i)
        start_index = min(dummy_stress_index) - 2
for i, s in enumerate(f06):
    if s.find('S T R E S S E S   I N   R O D   E L E M E N T S      ( C R O D )') != -1:
        dummy_stress_index2.append(i)
        end_index = min(dummy_stress_index2) - 4
stress = f06[start_index:end_index]
# ④ 응력 데이터 정리
del_list = []
for i, s in enumerate(stress):
    if s.find('S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M ') != -1:
        del_list.append(i)
del_list.reverse()
for j in range(0, len(del_list)):
    if j < len(del_list) - 1:
        del stress[del_list[j] - 4:del_list[j] + 3]
    if j == len(del_list) - 1:
        del stress[0:del_list[j] + 3]
for j in range(0, len(stress)):
    stress[j] = stress[j].split()
stress_final_list = []
for k in range(0, int(len(stress) / 3)):
    stress_final_list.append([int(stress[3 * k][1]), int(stress[3 * k + 1][0]), float(stress[3 * k + 1][1]), float(stress[3 * k + 1][6]), float(stress[3 * k + 1][7])])
    stress_final_list.append([int(stress[3 * k][1]), int(stress[3 * k + 2][0]), float(stress[3 * k + 2][1]), float(stress[3 * k + 2][6]), float(stress[3 * k + 2][7])])
df_stress = pd.DataFrame(stress_final_list, columns=['Elem_No', 'NODE', 'LENGTH', 'MAX', 'MIN'])
plus_stress = df_stress['MAX'].max()
minus_stress = abs(df_stress['MIN'].min())
stress_index = "MAX" if plus_stress > minus_stress else "MIN"
max_stress = plus_stress if stress_index == "MAX" else minus_stress
max_stress_element_no = df_stress.loc[df_stress['MAX'].idxmax(), 'Elem_No'] if stress_index == "MAX" else df_stress.loc[df_stress['MIN'].idxmin(), 'Elem_No']
max_stress_plot_no = int(df_stress.loc[df_stress['MAX'].idxmax(), 'LENGTH'] + 1) if stress_index == "MAX" else int(df_stress.loc[df_stress['MIN'].idxmin(), 'LENGTH'] + 1)

# ⑤ 장력에 해당 하는  Index 추출
tension_start_index = next((i for i, line in enumerate(f06) if "F O R C E S   I N   R O D   E L E M E N T S" in line), None)
tension_end_index = next((i for i, line in enumerate(f06) if " S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M )" in line), None)
tension = f06[tension_start_index + 3:tension_end_index - 4]

# ⑥ 응력 데이터 정리
cleaned_tension = []
for entry in tension:
    entry = entry.replace('\n', '').split()
    if len(entry) != 3:
        cleaned_tension.append([entry[0], entry[1]])
        cleaned_tension.append([entry[3], entry[4]])
    else:
        cleaned_tension.append([entry[0], entry[1]])
    df_tension = pd.DataFrame(cleaned_tension).iloc[:, [0, 1]]
    df_tension.columns = ["Elem_No", "TENSION[N]"]
    df_tension = df_tension.sort_values(by=['Elem_No'], ascending=[True], ignore_index=True)  # 정렬
    df_tension['TENSION[N]'] = [round(float(df_tension.iloc[i, 1]), 1) for i in range(0, len(df_tension))]

df_tension['Elem_No']=df_tension['Elem_No'].astype(int)

df_Set = df_Set.merge(df_tension[['Elem_No', 'TENSION[N]']], on='Elem_No', how='left')
df_Set["Name"] = df_Set.index + 1
df_Set["Name"]=df_Set["Name"].astype(str)
df_Set["Safety_factor"]=round((df_Set["BL[ton]"]/(df_Set["TENSION[N]"]/9800)),1)


# 5) hw 실행  ──────────────────────────────────────────────────────────────────────────────
# 5-1) hwc 파일 생성
hwc_total_path = Analysis_bdf.replace(".bdf", ".hwc")
with open(hwc_total_path, 'w') as f:
    f.writelines("open animation modelandresult " + '"' + Analysis_bdf + '"' + ' "' + Analysis_op2 + '"')
    f.writelines("\nresult scalar load type=Displacement\n")
    f.writelines("animate frame last\n")
    f.writelines("show legends\n")
    f.writelines("annotation note " + '"' + "Model Info" + '"' + "display visibility=false\n")
    f.writelines("scale deformed resulttype=Displacement value=10.000000\n")
    f.writelines("result scalar legend layout format=fixed\n")
    f.writelines("result scalar legend layout precision=1\n")
    f.writelines("result scalar legend values valuefont=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("result scalar legend title font=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  id=false\n")
    f.writelines("annotation measure global precision=1\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=false\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_disp.png") + '"')
    # ────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────
    if stress_index == "MAX":
        f.writelines("\nscale deformed resulttype=Displacement value=1.000000\n")
        f.writelines("result scalar load type=" + '"' + "1D Stress" + '"' + " component=" + '"' + "CBEAM Maximum Stress" + str(max_stress_plot_no) + '"' + "\n")
        f.writelines("annotation measure global precision=1\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=false\n")
    if stress_index == "MIN":
        f.writelines("\nscale deformed resulttype=Displacement value=1.000000\n")
        f.writelines("result scalar load type=" + '"' + "1D Stress" + '"' + " component=" + '"' + "CBEAM Minimum Stress" + str(max_stress_plot_no) + '"' + "\n")
        f.writelines("annotation measure global precision=1\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=True\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  max=false\n")
    f.writelines("animate frame last\n")
    f.writelines("show legends\n")
    f.writelines("result scalar legend layout format=fixed\n")
    f.writelines("result scalar legend layout precision=1\n")
    f.writelines("result scalar legend values valuefont=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("result scalar legend title font=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_stress.png") + '"' + "\n")
    # ────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────
    f.writelines("result scalar load type=" + '"' + "1D Force" + '"' + " component=" + '"' + "CROD Axial Force" + '"' + " system=global" + "\n")
    f.writelines("result scalar multiplier 0.000102041\n")
    for i, row in df_Set.iterrows():
        f.writelines("annotation measure create distancebetween " + '"' + "Measure Group 3" + '"' + "\n")
        f.writelines("annotation measure rename " + '"' + "Measure Group 3" + '"' + f" {row['Name']}" + "\n")
        f.writelines(f"annotation measure {row['Name']} type elementalcontour\n")
        f.writelines(f"annotation measure {row['Name']} add elements " + '"' + f"1 Beam {row['Elem_No']}" + '"' + "\n")
        f.writelines(f"annotation measure {row['Name']} display name=true\n")
        f.writelines("annotation measure global precision=0\n")
        f.writelines(f"annotation measure {row['Name']} display value=false\n")
        f.writelines(f"annotation measure {row['Name']} display id=false\n")
        f.writelines("annotation measure global transparency=true\n")
        f.writelines("annotation measure " + '"' +row['Name'] + '"' + " display font=" + '"' + "{Noto Sans} 8 normal roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '"' + "display visibility=false" + "\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_tension.png") + '"')

# 5-2) hw 파일 실행
hw = r"C:\Program Files\Altair\2022\hwdesktop\hw\bin\win64\hw.exe"

HWCScript = hwc_total_path
arg = hw + " -b -hwc " + HWCScript

subprocess.Popen(arg).wait()
img_disp_path = Analysis_bdf.replace(".bdf", "_disp.png")
img_stress_path = Analysis_bdf.replace(".bdf", "_stress.png")
img_tension_path = Analysis_bdf.replace(".bdf", "_tension.png")
while not (os.path.exists(img_disp_path) and os.path.exists(img_stress_path) and os.path.exists(img_tension_path)):
    missing_files = [file for file in [img_disp_path, img_stress_path, img_tension_path] if not os.path.exists(file)]
    if not missing_files:
        break  # 모든 파일이 생성될 경우 Break
    time.sleep(10)  # 1초 간격으로 확인

# 6) 해석 보고서 작성 ──────────────────────────────────────────────────────────────────────────────
Result_xlsx = file_path.replace(".bdf",".xlsx")
# 6-1) 보고서 양식 ─────────────────────────────────────────────────────────────────────────────────
# 1️⃣ 새 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# 2️⃣ 열 너비 설정
for col in ['A', 'G', 'H', 'N']:
    ws.column_dimensions[col].width = 14.5
for col in ['B', 'C', 'D', 'E', 'F', 'I', 'J', 'K', 'L', 'M']:
    ws.column_dimensions[col].width = 16

# 3️⃣ 셀 병합
merge_ranges = ['A1:A2', 'G1:G2', 'H1:H2', 'N1:N2', 'B1:F2', 'I1:M2',
                'B37:F37', 'B61:F61', 'I36:M36']
for rng in merge_ranges:
    ws.merge_cells(rng)

# 4️⃣ 값 및 글꼴 설정 (헤더)
ws['A1'] = "현대중공업"
ws['A1'].font = Font(bold=True, color="0070C0", size=11)

ws['B1'] = "승하선 사다리 구조 안전성 평가"
ws['B1'].font = Font(bold=True, color="000000", size=16)

ws['G1'] = "선박의장연구실"
ws['G1'].font = Font(bold=True, color="000000", size=11)

ws['H1'] = "현대중공업"
ws['H1'].font = Font(bold=True, color="0070C0", size=11)

ws['I1'] = "승하선 사다리 구조 안전성 평가"
ws['I1'].font = Font(bold=True, color="000000", size=16)

ws['N1'] = "선박의장연구실"
ws['N1'].font = Font(bold=True, color="000000", size=11)

# 5️⃣ 배경 흰색 적용
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
for row in ws["A1":"N61"]:
    for cell in row:
        cell.fill = white_fill

# 6️⃣ 병합 포함 지정 셀 중앙 정렬 + 얇은 테두리
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
center_align = Alignment(horizontal='center', vertical='center')

merged_cells = ['A1:A2', 'B1:F2', 'G1:G2', 'H1:H2', 'I1:M2', 'N1:N2',
                'B37:F37', 'B61:F61', 'I36:M36']
for rng in merged_cells:
    start_cell, end_cell = rng.split(':')
    min_row, max_row = ws[start_cell].row, ws[end_cell].row
    min_col, max_col = ws[start_cell].column, ws[end_cell].column
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border

# 7️⃣ 겉 테두리 영역 적용
outer_ranges = ['A3:G61', 'H3:N61']
for rng in outer_ranges:
    start_cell, end_cell = rng.split(':')
    min_row, max_row = ws[start_cell].row, ws[end_cell].row
    min_col, max_col = ws[start_cell].column, ws[end_cell].column
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            left = thin_side if col == min_col else None
            right = thin_side if col == max_col else None
            top = thin_side if row == min_row else None
            bottom = thin_side if row == max_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# 8️⃣ 화면 배율 55%
ws.sheet_view.zoomScale = 55

# 9️⃣ 추가 내용 적용

# B4
ws['B4'] = "1) 승하선용 사다리 모델 정보"
ws['B4'].font = Font(bold=True, color="548235", size=11)

# B5~C8: 글꼴, 정렬, 테두리
for row in range(5, 9):
    for col in [2, 3]:  # B=2, C=3
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="000000", size=11)
        cell.alignment = center_align
        cell.border = thin_border

# B5~C5: 연한 연두색 배경
light_green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
for col in [2, 3]:
    ws.cell(row=5, column=col).fill = light_green_fill
# B5
ws['C5'] = "승선 하중 [kg]"

# B10
ws['B10'] = "2) 승하선용 사다리 구조 강도 평가"
ws['B10'].font = Font(bold=True, color="548235", size=11)

# B11~F11
headers = ["재질", "인장 강도[MPa]", "최대 응력[MPa]", "안전율", "비고*"]
for i, val in enumerate(headers):
    col = 2 + i  # B=2
    ws.cell(row=11, column=col, value=val).font = Font(bold=True, color="000000", size=11)
    ws.cell(row=11, column=col).alignment = center_align
    ws.cell(row=11, column=col).border = thin_border
    ws.cell(row=11, column=col).fill = light_green_fill

# B12~F12: 테두리 + 중앙 정렬
for col in range(2, 7):  # B=2~F=6
    ws.cell(row=12, column=col).font = Font(bold=True, color="000000", size=11)
    ws.cell(row=12, column=col).alignment = center_align
    ws.cell(row=12, column=col).border = thin_border

# B12, C12
ws['B12'] ="SS400"
ws['C12'] ="400"

# B13 안내문
ws['B13'] = "* 인장 강도 기준으로 안전률이 3 이상인 경우: 구조 강도 만족"
ws['B13'].font = Font(color="000000", size=10)

# B37~F37 병합 및 값 입력
ws.merge_cells('B37:F37')
ws['B37'] = "[변형]"
ws['B37'].font = Font(bold=True, color="000000", size=11)
ws['B37'].alignment = center_align

# B61~F61 병합 및 값 입력
ws.merge_cells('B61:F61')
ws['B61'] = "[응력]"
ws['B61'].font = Font(bold=True, color="000000", size=11)
ws['B61'].alignment = center_align

# I36~M36 병합 및 값 입력
ws.merge_cells('I36:M36')
ws['I36'] = "[장력]"
ws['I36'].font = Font(bold=True, color="000000", size=11)
ws['I36'].alignment = center_align

# 10️⃣ 와이어 로프 강도 평가 추가
# I38: 제목
ws['I38'] = "3) 와이어 로프 강도 평가"
ws['I38'].font = Font(bold=True, color="548235", size=11)

# I39~M39: 헤더
headers_wire = ["위치", "와이어 직경", "계산 하중[ton]", "파단 하중[ton]", "안전율"]
for i, val in enumerate(headers_wire):
    col = 9 + i  # I=9
    ws.cell(row=39, column=col, value=val).font = Font(bold=True, color="000000", size=11)
    ws.cell(row=39, column=col).alignment = center_align
    ws.cell(row=39, column=col).border = thin_border
    ws.cell(row=39, column=col).fill = light_green_fill  # 연한 연두색 배경

# I40~M48: 내용 영역 테두리 + 중앙 정렬
for row in range(40, 49):
    for col in range(9, 14):  # I~M
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="000000", size=11)
        cell.alignment = center_align
        cell.border = thin_border

# I49~I50: 안내문
ws['I49'] = "* 안전율: 파단 하중 / 계산 하중"
ws['I50'] = "*  - 산업안전기준에 관한 규칙, 달기 와이어로프 및 달기 강선의 안전 계수: 10이상"
ws['I49'].font = Font(color="000000", size=10)
ws['I50'].font = Font(color="000000", size=10)

# 6-2) 해석 결과 입력 ────────────────────────────────────────────────────────────────────────────────────────
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# 1️⃣ 승하선용 사다리 모델 정보 입력
start_row=6
for i,row in df_Info.iterrows():
    ws[f"B{start_row + i}"] = row["Comp"]
    ws[f"C{start_row + i}"] = round(row["Load"]*100,1)
# 2️⃣ 응력 및 안전률 입력
ws['D12'] = round(max_stress,1)
ws['E12'] = round(400/max_stress,1)
if 400/max_stress < 3:
    ws['F12'].fill = yellow_fill
    ws['F12'] = "불만족"
    ws['F12'].font = Font(color="FF0000",bold=True)  # 빨간색
else:
    ws['F12'] = "만족"
    ws['F12'].font = Font(color="0000FF",bold=True)  # 파란색
# 3️⃣  와이어 로프 강도 평가 정보 입력
start_row=40
for i, row in df_Set.iterrows():
    ws[f"I{start_row + i}"] = row["Name"]
    ws[f"J{start_row + i}"] = row["Dia"]
    ws[f"K{start_row + i}"] = round(row["TENSION[N]"]/9800,1)
    ws[f"L{start_row + i}"] = row["BL[ton]"]
    cell = ws[f"M{start_row + i}"]
    cell.value = row["Safety_factor"]

    if row["Safety_factor"] > 10.0:
        cell.font = Font(color="0000FF", bold=True)  # 파란색 + Bold
    else:
        cell.fill = yellow_fill
        cell.font = Font(color="FF0000", bold=True)  # 빨간색 + Bold

# 4️⃣ 이미지 삽입
img_disp = Image(img_disp_path)
img_disp.width *= 0.645
img_disp.height *= 0.645
ws.add_image(img_disp, 'B14')
img_stress = Image(img_stress_path)
img_stress.width *= 0.645
img_stress.height *= 0.645
ws.add_image(img_stress, 'B38')
img_tension = Image(img_tension_path)
img_tension.width *= 0.645
img_tension.height *= 0.645
ws.add_image(img_tension, 'I8')
wb.save(Result_xlsx)
# 7) 관련 없는 파일 삭제 ──────────────────────────────────────────────────────────────────────────────
ext_list = [".hwc", ".log", ".f04", ".op2", "_disp.png", "_stress.png", "_tension.png"]
base_name = Analysis_bdf.replace(".bdf", "")
# 각 파일 삭제 시도
for ext in ext_list:
    del_file = base_name + ext
    if os.path.exists(del_file):
        os.remove(del_file)