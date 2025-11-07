import socket
import threading
import os
import subprocess
import time
import math
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Font
import sys
import shutil
import win32com.client as win32

#
# if len(sys.argv) < 2:
#     print("Usage: infoget_ladder.py <파일경로>")
#     sys.exit(1)

file_path = sys.argv[1]  # 첫 번째 인자: 파일 경로
# file_path = r"C:\Users\HHI\KHM\HiTessCloud_Flask\main\userConnection\HiTessBeam\LadderAnaysis\10.14.42.71_1755702901165\SUM2.bdf"  # 첫 번째 인자: 파일 경로
file_name = os.path.basename(file_path)
dir_path = os.path.dirname(file_path)

# 2) 해석 폴더 생성 ──────────────────────────────────────────────────────────────────────────────

Analysis_bdf = file_path
Analysis_f06 = file_path.replace(".bdf",".f06")
Analysis_op2 = file_path.replace(".bdf",".op2")

# 3) 해석 파일 실행 ──────────────────────────────────────────────────────────────────────────────

os.chdir(dir_path)
subprocess.Popen('nastran ' + Analysis_bdf).wait()


# 4) 해석 파일 분석 ──────────────────────────────────────────────────────────────────────────────
with open(Analysis_bdf, 'r', encoding='utf8') as f:
    bdf = f.readlines()
with open(Analysis_f06, 'r', encoding='utf8') as f:
    f06 = f.readlines()

# 4-1) bdf 분석
'''와이어 안전 하중 List 생성'''
list_BL = [[8.0, 3.5],
           [10.0, 5.6],
           [12.0, 8.4],
           [14.0, 11.2],
           [16.0, 14.7],
           [18.0, 18.9],
           [20.0, 23.1],
           [22.0, 28.0],
           [24.0, 33.6]]
'''내용 List 생성'''
list_Info=[]
list_Set = []
list_Prop = []
list_Rod = []
collecting = False
temp_set = []
## BDF 파일에서 내용 추출 (SET, Prop, Rod)
for i, s in enumerate(bdf):
    # 5.1 BDF 파일에서 내용 추출 (SET, MAT, PROP)
    if s.find('$-') == 0:  # '$-'가 문자열의 시작
        name, value = s.replace("$-", "").split(",")
        name_parts = name.rsplit("_", 1)[0]
        parts = name_parts.split("_")
        comp_name = "_".join(parts[:-1])
        length = parts[-1]
        if "p" in length:
            length = length.replace("p", ".")
        name_clean = f"{comp_name}_{length}"
        list_Info.append([name_clean, int(value)])  # value를 숫자로 변환
    if s.strip().startswith("SET"):
        collecting = True
        # "SET 1 =" 잘라내고 오른쪽 값만
        parts = s.split("=", 1)[1]
        nums = parts.replace(",", " ").split()
        temp_set.extend(int(x) for x in nums)
        if not s.strip().endswith(","):  # ,로 안 끝나면 SET 끝
            list_Set.append(temp_set)
            temp_set = []
            collecting = False
    # SET 이어지는 라인
    elif collecting:
        nums = s.replace(",", " ").split()
        temp_set.extend(int(x) for x in nums)
        if not s.strip().endswith(","):
            list_Set.append(temp_set)
            temp_set = []
            collecting = False
    if s.find('PROD ') != -1:
        Prop_No = s[8:16].strip()
        Mat_No = s[16:24].strip()
        Area = float(s[24:32].strip())
        Radius=math.sqrt(Area/math.pi)
        diameter = round(Radius, 1)*2
        prop_data = [int(Prop_No), int(Mat_No), float(Area),float(diameter)]
        list_Prop.append(prop_data)
    if s.find('CROD ') != -1:
        Elem_No = s[8:16].strip()
        Prop_No = s[16:24].strip()
        Node1 = s[24:32].strip()
        Node2 = s[32:40].strip()
        rod_data = [int(Elem_No), int(Prop_No), int(Node1), int(Node2)]
        list_Rod.append(rod_data)

'''Dataframe 생성'''
df_Set = pd.DataFrame(list_Set[0], columns=["Elem_No"])
df_Prop = pd.DataFrame(list_Prop, columns=['Prop_No', 'Mat_No','Area','Dia'])
df_Rod = pd.DataFrame(list_Rod, columns=['Elem_No', 'Prop_No', 'Node1', 'Node2'])
df_Info = pd.DataFrame(list_Info, columns=["Comp", "Load"])
df_BL = pd.DataFrame(list_BL, columns=["Dia", "BL[ton]"])
df_Set = df_Set.merge(df_Rod[['Elem_No', 'Prop_No']], on='Elem_No', how='left')
df_Set = df_Set.merge(df_Prop[['Prop_No','Dia']], on='Prop_No', how='left')
df_Set = df_Set.merge(df_BL[['Dia', 'BL[ton]']], on='Dia', how='left')

# 4-2) f06 분석
# ① 변형에 해당 하는 Index 추출
dummy_disp_index, dummy_disp_index2 = [], []
for i, s in enumerate(f06):
    if s.find('D I S P L A C E M E N T   V E C T O R') != -1:
        dummy_disp_index.append(i)
        start_disp_index = min(dummy_disp_index) - 2
for i, s in enumerate(f06):
    if s.find('F O R C E S   I N') != -1:
        dummy_disp_index2.append(i)
        end_disp_index = min(dummy_disp_index2) - 4
displacement = f06[start_disp_index:end_disp_index]
# ② 변형 데이터 정리
del_disp_list = []
for i, s in enumerate(displacement):
    if s.find('D I S P L A C E M E N T   V E C T') != -1:
        del_disp_list.append(i)
del_disp_list.reverse()
for j in range(0, len(del_disp_list)):
    if j < len(del_disp_list) - 1:
        del displacement[del_disp_list[j] - 4:del_disp_list[j] + 3]
    if j == len(del_disp_list) - 1:
        del displacement[0:del_disp_list[j] + 3]
for j in range(0, len(displacement)):
    displacement[j] = displacement[j].split()
disp_final_list = []
for k in range(0, len(displacement) - 1):
    disp_final_list.append([int(displacement[k][0]), float(displacement[k][2]), float(displacement[k][3]), float(displacement[k][4]), (float(displacement[k][2]) ** 2 + float(displacement[k][3]) ** 2 + float(displacement[k][4]) ** 2) ** (1 / 2)])
df_disp = pd.DataFrame(disp_final_list, columns=['NODE', 'DISP_X', 'DISP_Y', 'DISP_Z', 'DISP'])
max_disp = df_disp['DISP'].max()
max_disp_element_no = df_disp.loc[df_disp['DISP'].idxmax(), 'NODE']

# ③ 응력에 해당 하는  Index 추출
dummy_stress_index, dummy_stress_index2 = [], []
for i, s in enumerate(f06):
    if s.find('S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M ') != -1:
        dummy_stress_index.append(i)
        start_index = min(dummy_stress_index) - 2
for i, s in enumerate(f06):
    if s.find('S T R E S S E S   I N   R O D   E L E M E N T S      ( C R O D )') != -1:
        dummy_stress_index2.append(i)
        end_index = min(dummy_stress_index2) - 4
stress = f06[start_index:end_index]
# ④ 응력 데이터 정리
del_list = []
for i, s in enumerate(stress):
    if s.find('S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M ') != -1:
        del_list.append(i)
del_list.reverse()
for j in range(0, len(del_list)):
    if j < len(del_list) - 1:
        del stress[del_list[j] - 4:del_list[j] + 3]
    if j == len(del_list) - 1:
        del stress[0:del_list[j] + 3]
for j in range(0, len(stress)):
    stress[j] = stress[j].split()
stress_final_list = []
for k in range(0, int(len(stress) / 3)):
    stress_final_list.append([int(stress[3 * k][1]), int(stress[3 * k + 1][0]), float(stress[3 * k + 1][1]), float(stress[3 * k + 1][6]), float(stress[3 * k + 1][7])])
    stress_final_list.append([int(stress[3 * k][1]), int(stress[3 * k + 2][0]), float(stress[3 * k + 2][1]), float(stress[3 * k + 2][6]), float(stress[3 * k + 2][7])])
df_stress = pd.DataFrame(stress_final_list, columns=['Elem_No', 'NODE', 'LENGTH', 'MAX', 'MIN'])
plus_stress = df_stress['MAX'].max()
minus_stress = abs(df_stress['MIN'].min())
stress_index = "MAX" if plus_stress > minus_stress else "MIN"
max_stress = plus_stress if stress_index == "MAX" else minus_stress
max_stress_element_no = df_stress.loc[df_stress['MAX'].idxmax(), 'Elem_No'] if stress_index == "MAX" else df_stress.loc[df_stress['MIN'].idxmin(), 'Elem_No']
max_stress_plot_no = int(df_stress.loc[df_stress['MAX'].idxmax(), 'LENGTH'] + 1) if stress_index == "MAX" else int(df_stress.loc[df_stress['MIN'].idxmin(), 'LENGTH'] + 1)

# ⑤ 장력에 해당 하는  Index 추출
tension_start_index = next((i for i, line in enumerate(f06) if "F O R C E S   I N   R O D   E L E M E N T S" in line), None)
tension_end_index = next((i for i, line in enumerate(f06) if " S T R E S S E S   I N   B E A M   E L E M E N T S        ( C B E A M )" in line), None)
tension = f06[tension_start_index + 3:tension_end_index - 4]

# ⑥ 응력 데이터 정리
cleaned_tension = []
for entry in tension:
    entry = entry.replace('\n', '').split()
    if len(entry) != 3:
        cleaned_tension.append([entry[0], entry[1]])
        cleaned_tension.append([entry[3], entry[4]])
    else:
        cleaned_tension.append([entry[0], entry[1]])
    df_tension = pd.DataFrame(cleaned_tension).iloc[:, [0, 1]]
    df_tension.columns = ["Elem_No", "TENSION[N]"]
    df_tension = df_tension.sort_values(by=['Elem_No'], ascending=[True], ignore_index=True)  # 정렬
    df_tension['TENSION[N]'] = [round(float(df_tension.iloc[i, 1]), 1) for i in range(0, len(df_tension))]

df_tension['Elem_No']=df_tension['Elem_No'].astype(int)

df_Set = df_Set.merge(df_tension[['Elem_No', 'TENSION[N]']], on='Elem_No', how='left')
df_Set["Name"] = df_Set.index + 1
df_Set["Name"]=df_Set["Name"].astype(str)
df_Set["Safety_factor"]=round((df_Set["BL[ton]"]/(df_Set["TENSION[N]"]/9800)),1)


# 5) hw 실행  ──────────────────────────────────────────────────────────────────────────────
# 5-1) hwc 파일 생성
hwc_total_path = Analysis_bdf.replace(".bdf", ".hwc")
with open(hwc_total_path, 'w') as f:
    f.writelines("open animation modelandresult " + '"' + Analysis_bdf + '"' + ' "' + Analysis_op2 + '"')
    f.writelines("\nresult scalar load type=Displacement\n")
    f.writelines("animate frame last\n")
    f.writelines("show legends\n")
    f.writelines("annotation note " + '"' + "Model Info" + '"' + "display visibility=false\n")
    f.writelines("scale deformed resulttype=Displacement value=10.000000\n")
    f.writelines("result scalar legend layout format=fixed\n")
    f.writelines("result scalar legend layout precision=1\n")
    f.writelines("result scalar legend values valuefont=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("result scalar legend title font=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  id=false\n")
    f.writelines("annotation measure global precision=1\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=false\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_disp.png") + '"')
    # ────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────
    if stress_index == "MAX":
        f.writelines("\nscale deformed resulttype=Displacement value=1.000000\n")
        f.writelines("result scalar load type=" + '"' + "1D Stress" + '"' + " component=" + '"' + "CBEAM Maximum Stress" + str(max_stress_plot_no) + '"' + "\n")
        f.writelines("annotation measure global precision=1\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=false\n")
    if stress_index == "MIN":
        f.writelines("\nscale deformed resulttype=Displacement value=1.000000\n")
        f.writelines("result scalar load type=" + '"' + "1D Stress" + '"' + " component=" + '"' + "CBEAM Minimum Stress" + str(max_stress_plot_no) + '"' + "\n")
        f.writelines("annotation measure global precision=1\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display visibility=true\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display font=" + '"' + "{Noto Sans} 14 bold roman" + '"' + "\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  min=True\n")
        f.writelines("annotation measure " + '"' + "Static MinMax Result" + '" ' + "display  max=false\n")
    f.writelines("animate frame last\n")
    f.writelines("show legends\n")
    f.writelines("result scalar legend layout format=fixed\n")
    f.writelines("result scalar legend layout precision=1\n")
    f.writelines("result scalar legend values valuefont=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("result scalar legend title font=" + '"' + "{Noto Sans} 15 bold roman" + '"' + "\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_stress.png") + '"' + "\n")
    # ────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────
    f.writelines("result scalar load type=" + '"' + "1D Force" + '"' + " component=" + '"' + "CROD Axial Force" + '"' + " system=global" + "\n")
    f.writelines("result scalar multiplier 0.000102041\n")
    for i, row in df_Set.iterrows():
        f.writelines("annotation measure create distancebetween " + '"' + "Measure Group 3" + '"' + "\n")
        f.writelines("annotation measure rename " + '"' + "Measure Group 3" + '"' + f" {row['Name']}" + "\n")
        f.writelines(f"annotation measure {row['Name']} type elementalcontour\n")
        f.writelines(f"annotation measure {row['Name']} add elements " + '"' + f"1 Beam {row['Elem_No']}" + '"' + "\n")
        f.writelines(f"annotation measure {row['Name']} display name=true\n")
        f.writelines("annotation measure global precision=0\n")
        f.writelines(f"annotation measure {row['Name']} display value=false\n")
        f.writelines(f"annotation measure {row['Name']} display id=false\n")
        f.writelines("annotation measure global transparency=true\n")
        f.writelines("annotation measure " + '"' +row['Name'] + '"' + " display font=" + '"' + "{Noto Sans} 8 normal roman" + '"' + "\n")
    f.writelines("annotation measure " + '"' + "Static MinMax Result" + '"' + "display visibility=false" + "\n")
    f.writelines("save image page " + '"' + Analysis_bdf.replace(".bdf", "_tension.png") + '"')

# 5-2) hw 파일 실행
hw = r"C:\Program Files\Altair\2022\hwdesktop\hw\bin\win64\hw.exe"
HWCScript = hwc_total_path
arg = hw + " -b -hwc " + HWCScript

subprocess.Popen(arg).wait()
img_disp_path = Analysis_bdf.replace(".bdf", "_disp.png")
img_stress_path = Analysis_bdf.replace(".bdf", "_stress.png")
img_tension_path = Analysis_bdf.replace(".bdf", "_tension.png")
while not (os.path.exists(img_disp_path) and os.path.exists(img_stress_path) and os.path.exists(img_tension_path)):
    missing_files = [file for file in [img_disp_path, img_stress_path, img_tension_path] if not os.path.exists(file)]
    if not missing_files:
        break  # 모든 파일이 생성될 경우 Break
    time.sleep(10)  # 1초 간격으로 확인

# 6) 해석 보고서 작성 ──────────────────────────────────────────────────────────────────────────────
original_file_path = r"C:\Users\HHI\KHM\HiTessCloud_Flask\main\PythonModule\Reference\Report_Ladder.xlsx"
Result_xlsx = file_path.replace(".bdf",".xlsx")

excel = win32.Dispatch("Excel.Application")
excel.DisplayAlerts = False
wb = excel.Workbooks.Open(original_file_path, ReadOnly=True)    # 필요시 비번 있으면 Password:="..."
wb.SaveAs(Result_xlsx, FileFormat=51)
wb.Close(False)
excel.Quit()

# 2) openpyxl로 열기
wb2 = load_workbook(Result_xlsx, data_only=True)
ws = wb2.active
print(ws["B12"].value)

# with open(original_file_path, 'rb') as f:
#     with open(Result_xlsx, 'wb') as new_f:
#         new_f.write(f.read())
#
# yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# print("1차 진행")
# # ## 값 변경
# wb = load_workbook(filename=original_file_path)
# sheet = wb["구조 강도 평가"]
# print("2차 진행")
# # 6-1) 승하선용 사다리 모델 정보 입력
# start_row=6
# for i,row in df_Info.iterrows():
#     sheet[f"B{start_row + i}"] = row["Comp"]
#     sheet[f"C{start_row + i}"] = round(row["Load"]*100,1)
#
# # 6-2) 응력 및 안전률 입력
# sheet['D12'] = max_stress
# sheet['E12'] = 400/max_stress
# if 400/max_stress < 3:
#     sheet['F12'].fill = yellow_fill
#     sheet['F12'] = "불만족"
#     sheet['F12'].font = Font(color="FF0000",bold=True)  # 빨간색
# else:
#     sheet['F12'] = "만족"
#     sheet['F12'].font = Font(color="0000FF",bold=True)  # 파란색
# # 6-4) 와이어 로프 강도 평가 정보 입력
# start_row=40
# for i, row in df_Set.iterrows():
#     sheet[f"I{start_row + i}"] = row["Name"]
#     sheet[f"J{start_row + i}"] = row["Dia"]
#     sheet[f"K{start_row + i}"] = row["TENSION[N]"]/9800
#     sheet[f"L{start_row + i}"] = row["BL[ton]"]
#     cell = sheet[f"M{start_row + i}"]
#     cell.value = row["Safety_factor"]
#
#     if row["Safety_factor"] > 10.0:
#         cell.font = Font(color="0000FF", bold=True)  # 파란색 + Bold
#     else:
#         cell.fill = yellow_fill
#         cell.font = Font(color="FF0000", bold=True)  # 빨간색 + Bold
#
# # 6-5) 이미지 삽입
# img_disp = Image(img_disp_path)
# img_disp.width *= 0.645
# img_disp.height *= 0.645
# sheet.add_image(img_disp, 'B14')
# img_stress = Image(img_stress_path)
# img_stress.width *= 0.645
# img_stress.height *= 0.645
# sheet.add_image(img_stress, 'B38')
# img_tension = Image(img_tension_path)
# img_tension.width *= 0.645
# img_tension.height *= 0.645
# sheet.add_image(img_tension, 'I8')
# wb.save(Result_xlsx)
# # 7) 관련 없는 파일 삭제 ──────────────────────────────────────────────────────────────────────────────
# ext_list = [".hwc", ".log", ".f04", ".op2", "_disp.png", "_stress.png", "_tension.png"]
# base_name = Analysis_bdf.replace(".bdf", "")
# # 각 파일 삭제 시도
# for ext in ext_list:
#     del_file = base_name + ext
#     if os.path.exists(del_file):
#         os.remove(del_file)